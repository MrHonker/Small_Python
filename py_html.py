#!/usr/bin/env python
# -*- coding: utf-8 -*-

import sys
import time
from smtplib import SMTP
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import datetime
import os.path  
import mimetypes
# python2.x
# from email.MIMEBase import MIMEBase
from email.mime.base import MIMEBase
# from email.encode import Encoders 

import pickle
from openpyxl import Workbook
from openpyxl import load_workbook
# from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import ExcelWriter 
from email.header import Header  
import smtplib    


def send_mail(config):
    print("Sending Mail...")
    smtpserver = 'smtp.163.com'
    username = '13407178083@163.com'
    password='20090607h'
    sender='13407178083@163.com'
    #receiver='XXX@126.com'
    #收件人为多个收件人
    subject = 'this is excel email'
    receiver=['13407178083@163.com']
    msg = MIMEMultipart('mixed') 
    msg['Subject'] = Header(subject, 'utf-8')
    msg['From'] = '13407178083@163.com'
    msg['To'] = '13407178083@163.com'
    text = config['text']
    msg.attach(text)
    # 登录并发送邮件
    try:
        smtp = smtplib.SMTP()
        smtp.connect(smtpserver)
        smtp.login(username, password)
        smtp.sendmail(sender, receiver, msg.as_string())
    except:
        print("邮件发送失败！")
    else:
        print("邮件发送成功！")
    finally:
        smtp.quit()





def send_mail_to_test(context):
    send_mail({
        'to': ["13407178083@163.com"],
        'cc': ['13407178083@163.com'],
        'server': 'smtp.163.com',
        'port': 25,
        'subject': 'Just for Test',
        'username': '13407178083@163.com',
        'password': '20090607h',
        'text': context}
    )

def message_from_excel():
    wb = load_workbook(fileName,data_only=True)
    ws = wb.get_sheet_by_name('Crash')

    all_versions = []
    personNums = []
    hanppends = []

    today_bugly = []
    yes_bugly = []

    bugly_flu = []

    for rown in range(3,7):
        for coln in range(2,8):
            value = ws.cell(row=rown,column=coln).value
            if coln == 2:
                all_versions.append(value)
            elif coln == 3:
                personNums.append(int(value))
            elif coln == 4:
                hanppends.append(int(value))
            elif coln == 5:
                today_bugly.append(float(value))
            elif coln == 6:
                yes_bugly.append(float(value))
            

    for x in range(0,4):  
        bugly_flu.append(crash_rate(today_bugly[x],yes_bugly[x]))

    html = """\
<!DOCTYPE html>
<html>
<meta charset="utf-8">
<head>
    <title>iOS - Bugly崩溃日报</title>
</head>
<body>
<div id="container">
    <div id="content">
        <p>

            版本崩溃信息：
            <table width="800" border="2" bordercolor="black" cellspacing="2">
                <tr>
                    <td><strong>版本号</strong></td>
                    <td><strong>影响人数</strong></td>
                    <td><strong>发生次数</strong></td>
                    <td><strong>日崩溃率-用户指标</strong></td>
                    <td><strong>波动</strong></td>
                </tr>
                <tr>
                    <td>""" + str(all_versions[0]) + """</td>
                    <td>""" + str(personNums[0]) + """</td>
                    <td>""" + str(hanppends[0]) + """</td>
                    <td>""" + daily_crash_bugly(today_bugly[0]) + """</td>
                    <td bgcolor="#FF8040">""" + str(bugly_flu[0]) + """</td>
                </tr>
                <tr>
                    <td>""" + str(all_versions[1]) + """</td>
                    <td>""" + str(personNums[1]) + """</td>
                    <td>""" + str(hanppends[1]) + """</td>
                    <td>""" + daily_crash_bugly(today_bugly[1]) + """</td>
                    <td bgcolor="#FF8040">""" + str(bugly_flu[1]) + """</td>
                </tr>
                <tr>
                    <td>""" + str(all_versions[2]) + """</td>
                    <td>""" + str(personNums[2]) + """</td>
                    <td>""" + str(hanppends[2]) + """</td>
                    <td>""" + daily_crash_bugly(today_bugly[2]) + """</td>
                    <td bgcolor="#FF8040">""" + str(bugly_flu[2]) + """</td>
                </tr>
                <tr>
                    <td>""" + str(all_versions[3]) + """</td>
                    <td>""" + str(personNums[3]) + """</td>
                    <td>""" + str(hanppends[3]) + """</td>
                    <td>""" + daily_crash_bugly(today_bugly[3]) + """</td>
                    <td bgcolor="#FF8040">""" + str(bugly_flu[3]) + """</td>
                </tr>
            </table>
        </p>
        <p>

        详情请见附件

        </p>
    </div>
</div>
</body>
</html>
    """

    context = MIMEText(html,_subtype='html',_charset='utf-8')
    send_mail_to_test(context) 

def crash_rate(today,yester):  
    rate = "%.2f" %(float(today) - float(yester))
    return rate + '%'

def daily_crash_bugly(num):
    temp = "%.2f" %(num * 100) + '%'
    return temp

fileName = "E:/python_prj/python_email/auto_email/Small_Python/Bugly-Daily-iOS.xlsx"
print(fileName)

if __name__ == '__main__':
    message_from_excel()
